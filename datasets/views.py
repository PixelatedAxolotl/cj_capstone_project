import openpyxl
from django.db import transaction
import base64
import io
from datetime import datetime, timezone
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse


from .data_validators import validate_columns
from .constants import Q7_PARTICIPATION_COLS, SURVEY_COLUMNS, Q6_PARTICIPATION_COLS
from .models import Dataset, Question, Response, RespondentAnswer, QuestionColumn, Option
from core.models import School
from .data_access_control import get_dashboard_queryset, get_response_queryset, can_view_raw, get_aggregate_scopes
from .crosstab_builder import build_crosstab
from .visualizations import build_crosstab_table, build_combined_crosstab_table, build_grouped_bar, build_participation_chart, build_top_selections, build_aptitude_summary, build_career_cluster_top3, build_post_hs_conversations
from accounts.decorators import role_required

# TODO: stop dataset upload and check with user if too many duplicate response IDs are found
@role_required('internal')  # only allow admin users access
def upload_dataset(request):
    print("You are in upload_dataset!")
    context = {}

    if request.method == 'POST' and request.FILES.get('dataset_file'):
        file = request.FILES['dataset_file']

        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Only .xlsx files are supported.')
            return render(request, 'datasets/upload.html', context)

        try:
            # Read file contents and encode as base64 for session storage
            file_contents = file.read()
            request.session['upload_file_data'] = base64.b64encode(file_contents).decode('utf-8')
            request.session['upload_file_name'] = file.name

            # Parse for preview
            wb = openpyxl.load_workbook(io.BytesIO(file_contents), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            headers = list(rows[0])
            row_count = len(rows) - 1
            data_rows = rows[1:]
            wb.close()

            context = {
                'preview': True,
                'headers': headers,
                'row_count': row_count,
                'data_rows': data_rows[:50], # pass only first 50 rows to preview table FOR NOW - TODO: REMOVE THIS LIMIT LATER
                'validation': validate_columns(headers),
                'file_name': file.name,
            }

        except Exception as e:
            messages.error(request, f'Could not parse file: {str(e)}')

    return render(request, 'datasets/upload.html', context)



# TODO: speed better but add fun little loading animation?
#   make deleted datasets recoverable?
@role_required('internal')
def confirm_import(request):

    # Retrieve file data stored in session during upload_dataset step
    file_data = request.session.get('upload_file_data')
    file_name = request.session.get('upload_file_name', '')

    if not file_data:
        messages.error(request, 'No file found. Please upload a file first.')
        return redirect('upload_dataset')

    if request.method == 'POST':
        dataset_name = request.POST.get('dataset_name', '').strip() or file_name

        try:
            with transaction.atomic():  # ensure all-or-nothing import — no partial data is ever committed

                # Create the Dataset record first so responses can reference it
                dataset = Dataset.objects.create(
                    name=dataset_name,
                    description='',
                    row_count=0,
                )

                # Decode file from session and parse
                file_contents = base64.b64decode(file_data)
                wb = openpyxl.load_workbook(io.BytesIO(file_contents), read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(min_row=1, values_only=True))
                wb.close()

                headers = list(all_rows[0])
                data_rows = all_rows[1:]

                # Map header name to its column index for lookup during row processing
                header_index = {header: index for index, header in enumerate(headers) if header}

                # Pre-fetch all QuestionColumn records into memory to avoid per-row DB queries
                question_column_map = {
                    question_column.column_header: question_column
                    for question_column in QuestionColumn.objects.select_related('question', 'option').all()
                }

                # fetch schools keyed by survey_index to map Q1 numeric values to school names - done here to avoid per-row queries during import loop
                schools_by_index = {
                    school.survey_index: school
                    for school in School.objects.filter(survey_index__isnull=False)
                }

                # fetch options keyed by (category, numeric_value) for single choice and scale questionlookups - done here to avoid per-row queries during import loop
                option_lookup = {
                    (option.category, option.numeric_value): option
                    for option in Option.objects.filter(category__isnull=False)
                }

                def parse_value(raw, cast):
                    #Convert raw cell values to the appropriate Python type for DB insertion.
                    if raw is None:
                        return None
                    try:
                        if cast == 'datetime':
                            # Assumes Qualtrics export format — update if export format changes
                            # Timestamps stored as UTC
                            return datetime.strptime(raw, "%m/%d/%Y %H:%M:%S").replace(tzinfo=timezone.utc).isoformat()
                        elif cast == 'int':
                            return int(float(raw))
                        elif cast == 'float':
                            return float(raw)
                        elif cast == 'str':
                            return str(raw)
                    except (ValueError, TypeError):
                        return None

                imported_count = 0
                skipped_duplicates = []

                # Use set for faster lookups; grows during the loop to catch within-file duplicates too
                existing_response_ids = set(Response.objects.values_list('response_id', flat=True))

                responses_to_create = []
                all_answer_kwargs = []

                for raw_row in data_rows:
                    # Map each header to its cell value for this row
                    row_data = {headers[i]: raw_row[i] for i in range(len(headers)) if i < len(raw_row)}

                    # Skip rows with no ResponseId
                    response_id = row_data.get('ResponseId')
                    if not response_id or str(response_id).strip() == '':
                        continue

                    # Skip duplicate responses and warn user — ResponseId is unique per survey submission
                    response_id_str = str(response_id)
                    if response_id_str in existing_response_ids:
                        print(f"Skipping duplicate ResponseId {response_id} in row with data {row_data}")  # log duplicate for debugging
                        skipped_duplicates.append(response_id_str)
                        continue
                    existing_response_ids.add(response_id_str)  # prevent within-file duplicates too

                    # Build kwargs for Response and RespondentAnswer creation - use SURVEY_COLUMNS to determine field mapping and handling logic for each column
                    response_kwargs = {'dataset': dataset}
                    answer_kwargs_list = []

                    for column_header, column_config in SURVEY_COLUMNS.items():
                        if column_config.get('skip') or column_header not in header_index:
                            continue

                        raw_value = row_data.get(column_header)
                        field_type = column_config['field_type']

                        if field_type in ('metadata', 'calculated'):
                            # set defualts in case cast option isn't set in constants.py
                            cast = column_config.get('cast', 'float' if field_type == 'calculated' else 'str')
                            response_kwargs[column_config['model_field']] = parse_value(raw_value, cast)

                        elif field_type == 'school':
                            # Resolve school FK from Q1 numeric value via survey_index
                            if raw_value is not None:
                                try:
                                    school_model_field = schools_by_index.get(parse_value(raw_value, column_config.get('cast', 'int')))
                                    if school_model_field is not None:
                                        response_kwargs[column_config['model_field']] = school_model_field
                                    else:
                                        print(f"Warning: No school found with survey_index {raw_value} for row with ResponseId {response_id}")
                                except (ValueError, TypeError):
                                    pass

                        elif field_type == 'discard':
                            continue

                        elif field_type == 'answer':
                            # Store in RespondentAnswer — handling varies by question type
                            question_column = question_column_map.get(column_header)
                            if question_column is None:
                                continue
                            question_type = question_column.question.question_type

                            if question_type == 'binary':
                                # Only store if selected (value == 1)
                                if raw_value in (1, '1', 1.0):
                                    answer_kwargs_list.append({
                                        'question_column': question_column,
                                        'option': question_column.option,
                                    })

                            # For free text and rank questions, store raw text value/rank number in the text_value field
                            elif question_type == 'free_text' or question_type == 'rank':
                                if raw_value is not None and str(raw_value).strip():
                                    answer_kwargs_list.append({
                                        'question_column': question_column,
                                        'text_value': str(raw_value),
                                    })

                            elif question_type in ('single_choice', 'scale'):
                                # Look up Option by category and numeric value
                                if raw_value is not None:
                                    option = option_lookup.get((question_column.option_category, int(raw_value)))
                                    if option:
                                        answer_kwargs_list.append({
                                            'question_column': question_column,
                                            'option': option,
                                        })


                    # Derive participation fields from raw row data - see constants.py for constant list definitions
                    # if respondent answered yes to any of the participation questions in Q6, mark participated_awareness as yes, otherwise no
                    # if respondent answered yes to any of the participation questions in Q7, mark participated_exploration as yes, otherwise no
                    # if respondent answered yes to any of the participation questions in either Q6 or Q7, mark participated_either as yes, otherwise no
                    participated_q6 = any(row_data.get(col) in (1, '1', 1.0) for col in Q6_PARTICIPATION_COLS)
                    participated_q7 = any(row_data.get(col) in (1, '1', 1.0) for col in Q7_PARTICIPATION_COLS)

                    response_kwargs['particip_career_prep_awareness']   = True if participated_q6 else False
                    response_kwargs['particip_career_prep_exploration'] = True if participated_q7 else False
                    response_kwargs['particip_career_prep_either']      = True if (participated_q6 or participated_q7) else False


                    responses_to_create.append(Response(**response_kwargs))
                    all_answer_kwargs.append(answer_kwargs_list)

                # Bulk insert all responses in a single query - Django sets PKs on returned objects
                created_responses = Response.objects.bulk_create(responses_to_create)

                # Build and bulk insert all answers in a single query
                answer_objects = []
                for response, answer_kwargs_list in zip(created_responses, all_answer_kwargs):
                    for kwargs in answer_kwargs_list:
                        answer_objects.append(RespondentAnswer(response=response, **kwargs))
                if answer_objects:
                    RespondentAnswer.objects.bulk_create(answer_objects)

                imported_count = len(created_responses)

                # Update dataset row count now that all rows have been processed
                dataset.row_count = imported_count
                dataset.save()

            # Clear session file data on successful import
            del request.session['upload_file_data']
            del request.session['upload_file_name']

            if skipped_duplicates:
                messages.warning(
                    request,
                    f'Import complete. {imported_count} rows imported. '
                    f'{len(skipped_duplicates)} duplicate response IDs were skipped: '
                    f'{",".join(skipped_duplicates[:10])}'
                    f'{"..." if len(skipped_duplicates) > 10 else ""}. '
                    f'You may be uploading a dataset that has already been imported.'
                )
            else:
                messages.success(request, f'Import complete. {imported_count} rows imported successfully.')

            return redirect('upload_complete')

        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}. No data was saved.')
            return render(request, 'datasets/confirm_import.html', {
                'file_name': file_name,
                'dataset_name': file_name,
            })

    # GET — show confirmation form with filename pre-populated as dataset name
    return render(request, 'datasets/confirm_import.html', {
        'file_name': file_name,
        'dataset_name': file_name.replace('.xlsx', ''),
    })



@role_required('internal')
def upload_complete(request):
    return render(request, 'datasets/upload_complete.html')


@role_required('internal')
def dataset_detail(request, dataset_id):

    dataset = Dataset.objects.get(pk=dataset_id)

    # Handle dataset name/description edits
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if name:
            dataset.name = name
            dataset.description = description
            dataset.save()
            messages.success(request, 'Dataset updated successfully.')

    # Fetch all RespondentAnswer records for dataset
    # select_related prevents additional queries when accessing related fields.
    answers = (
        RespondentAnswer.objects
        .filter(response__in=get_response_queryset(request.user), response__dataset=dataset)
        .select_related('response', 'question_column', 'question_column__question', 'option')
    )

    # Construct dictionary keyed by response ID
    # For each answer, store either the free text value, the numeric option value,
    # or 1 for binary questions (which only create a record when selected).
    response_answers = defaultdict(dict)
    for respondent_answer in answers:

        column_header = respondent_answer.question_column.column_header

        if respondent_answer.text_value is not None:
            response_answers[respondent_answer.response.id][column_header] = respondent_answer.text_value

        elif respondent_answer.option and respondent_answer.option.numeric_value is not None:
            response_answers[respondent_answer.response.id][column_header] = respondent_answer.option.numeric_value

        else:
            response_answers[respondent_answer.response.id][column_header] = 1

    # Fetch Response records to get metadata, calculated score, and school fields
    # select_related used to avoid additional queries when accessing related school field
    if not can_view_raw(request.user):
        messages.error(request, 'You do not have permission to view raw response data.')
        return redirect('dashboard')

    responses = list(
        get_response_queryset(request.user)
        .filter(dataset=dataset)
        .select_related('school')
    )


    for response in responses:
        response_id = response.id

        for column_header, column_config in SURVEY_COLUMNS.items():
            field_type = column_config['field_type']

            if field_type in ('metadata', 'calculated') and 'skip' not in column_config:
                response_answers[response_id][column_header] = getattr(response, column_config['model_field'], '')

            elif field_type == 'school':
                response_answers[response_id][column_header] = response.school.name if response.school else ''

    # Build header list from SURVEY_COLUMNS, excluding discarded columns.
    # Column order is determined by the order of entries in SURVEY_COLUMNS.
    headers = [column_header for column_header, column_config in SURVEY_COLUMNS.items()
               if column_config['field_type'] != 'discard']

    # Build rows in column order for display in template.
    # Missing values default to empty string.
    data_rows = [
        [response_answers[response.id].get(column_header, '') for column_header in headers]
        for response in responses
    ]

    # pass dataset info, headers, and data rows to template for rendering
    context = {
        'dataset': dataset,
        'headers': headers,
        'data_rows': data_rows,
        'row_count': len(data_rows),
    }
    return render(request, 'datasets/dataset_detail.html', context)



# Renders the HTML shell
@login_required
def dashboard_view(request):
    """
        Render the summary dashboard view.
        The frontend makes separate AJAX calls to fetch data for each chart/table.
        args:
        returns:
            response with rendered html template for dashboard view
            additional context variables:
                groups: list of groups the user has access to (based on what exists in their allowed datasets) for filtering (id and name)
                years: list of years present in the allowed datasets for filtering
                schools: (admin users ONLY) list of schools the user has access to for filtering (id and name)
                is_global: boolean indicating whether user has global admin access (used to determine whether to show school filter)

    """
    user = request.user
    qs   = get_dashboard_queryset(user).distinct()

    groups = (
        qs.values('school__groups__id', 'school__groups__name')
        .exclude(school__groups__isnull=True)
        .distinct()
    )
    years = [d.year for d in qs.dates('start_date', 'year')]

    # Users with 'all' scope are internal admins — show list of schools instead of "My School"
    is_global = 'all' in get_aggregate_scopes(user)
    schools = (
        qs.values('school__id', 'school__name')
        .exclude(school__isnull=True)
        .distinct()
        .order_by('school__name')
    ) if is_global else []

    return render(request, 'datasets/summary_dashboard.html', {
        'groups':    groups,
        'years':     sorted(years, reverse=True),
        'schools':   schools,
        'is_global': is_global,
    })



@login_required
def crosstab_tables_view(request):
    user = request.user
    qs = get_dashboard_queryset(user).distinct()
    groups = (
        qs.values("school__groups__id", "school__groups__name")
        .exclude(school__groups__isnull=True)
        .distinct()
    )
    years = [d.year for d in qs.dates("start_date", "year")]
    questions = (
        Question.objects
        .filter(can_crosstab=True)
        .order_by('crosstab_label')
    )

    return render(request, "datasets/crosstabs.html", {
        "groups": groups,
        "years": sorted(years, reverse=True),
        "questions": questions,
        "view_mode": "tables",
    })

@login_required
def dashboard_data(request):
    qs = get_dashboard_queryset(request.user).distinct()

    mode = request.GET.get('mode', 'percentages')
    group = request.GET.get('group', 'all')
    year  = request.GET.get('year', 'all')

    if group != 'all':
        if group == 'my':
            qs = qs.filter(school=request.user.school)
        elif group.startswith('group_'):
            qs = qs.filter(school__groups__id=group.split('_')[1]).distinct()
        elif group.startswith('school_'):
            qs = qs.filter(school__id=group.split('_')[1])

    if year != 'all':
        qs = qs.filter(start_date__year=year)

    if not qs.exists():
        return JsonResponse({'empty': True, 'message': 'No student responses found.'})

    grade_q = QuestionColumn.objects.get(column_header='Q5')
    plan_q  = QuestionColumn.objects.get(column_header='Q14')

    # build data for future plans broken down by letter grade chart
    plans_results = build_crosstab(
        x_question_id=grade_q.question_id,
        base_queryset=qs,
        y_question_ids=[plan_q.question_id],
        pct_type='row',
    )

    # build data for career prep participation broken down by letter grade chart
    participation = build_participation_chart(qs, grade_q, mode)


    # build data for top interests and skills
    top_interests = build_top_selections(qs, col_prefix='Q3_', exclude_cols={'Q3_16', 'Q3_16_TEXT'})
    top_skills    = build_top_selections(qs, col_prefix='Q25R_')
    top_traits    = build_top_selections(qs, col_prefix='Q26A_')

    aptitude        = build_aptitude_summary(qs)
    # TODO: make order by #1 with grade
    # TODO: enforce #1 with grade can only be displayed once for each average grade
    # TODO: agg C and below into single group
    cluster_top3    = build_career_cluster_top3(qs)
    conversations   = build_post_hs_conversations(qs)

    if not plans_results or not participation:
        return JsonResponse({'empty': True, 'message': 'No data found for these questions.'})

    return JsonResponse({
        'empty':           False,
        'plans_vs_grade':  build_grouped_bar(plans_results[0], mode, title="Post High School Plans By Average Letter Grade",
                                             legend_title="Post High School Plans",
                                             x_axis_title="Average Letter Grade").to_dict(),
        'participation':   participation,
        'top_interests':   top_interests,
        'top_skills':      top_skills,
        'top_traits':      top_traits,
        'aptitude':        aptitude,
        'cluster_top3':    cluster_top3,
        'conversations':   conversations,
    })


@login_required
def crosstab_data(request):
    x_id  = request.GET.get('x')
    y_ids = request.GET.getlist('y')
    mode  = request.GET.get('mode', 'counts')
    group = request.GET.get('group', 'all')
    year  = request.GET.get('year', 'all')

    if not x_id or not y_ids:
        return JsonResponse({'error': 'x and at least one y are required'}, status=400)

    try:
        x_id  = int(x_id)
        y_ids = [int(i) for i in y_ids]
    except ValueError:
        return JsonResponse({'error': 'x and y must be integers'}, status=400)

    if mode not in ('counts', 'percentages'):
        return JsonResponse({'error': 'mode must be counts or percentages'}, status=400)

    qs = get_dashboard_queryset(request.user).distinct()

    if group != 'all':
        if group == 'my':
            qs = qs.filter(school=request.user.school)
        elif group.startswith('group_'):
            qs = qs.filter(school__groups__id=group.split('_')[1]).distinct()

    if year != 'all':
        qs = qs.filter(start_date__year=year)

    if not qs.exists():
        return JsonResponse({'empty': True, 'message': 'No student responses found.'})

    results = build_crosstab(x_question_id=x_id,
                             y_question_ids=y_ids,
                             base_queryset=qs,
                             pct_type='column')

    if not results:
        return JsonResponse({'empty': True, 'message': 'No data found for the selected questions.'})

    return JsonResponse({
        'empty':    False,
        'combined': build_combined_crosstab_table(results, mode).to_dict(),
        'individual': [
            {
                'table': build_crosstab_table(result, mode).to_dict(),
                'bar':   build_grouped_bar(result, mode).to_dict(),
            }
            for result in results
        ],
    })